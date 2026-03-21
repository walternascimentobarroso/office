# -*- coding: utf-8 -*-
"""Excel service for generating Excel files from JSON input"""

from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any, Optional, List
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging

from src.core.business_calendar import last_weekday_of_month
from src.core.config import Config
from src.core.mapping import MappingLoader
from src.core.exceptions import TemplateLoadError, ExcelWriteError
from src.models.request import GenerateExcelRequest, Meta, Entry
from src.services.date_service import DateService

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for loading templates and generating Excel files"""

    def __init__(self, config: Optional[Config] = None):
        """Initialize service with configuration"""
        self.config = config or Config()
        self.header_mapping = MappingLoader.load_header_mapping(
            self.config.header_mapping_path
        )
        self.rows_mapping = MappingLoader.load_rows_mapping(
            self.config.rows_mapping_path
        )
        self.footer_mapping = MappingLoader.load_footer_mapping(
            self.config.footer_mapping_path
        )
        logger.info(
            "ExcelService initialized",
            extra={
                "extra_data": {
                    "template_path": self.config.template_path,
                    "header_mapping_keys": list(self.header_mapping.keys()),
                    "rows_mapping_keys": list(self.rows_mapping["columns"].keys()),
                    "footer_mapping_keys": list(self.footer_mapping.keys()),
                }
            },
        )

    def load_template(self) -> Any:
        """Load Excel template from configured path"""
        try:
            template_path = Path(self.config.template_path)
            if not template_path.exists():
                raise TemplateLoadError(f"Template not found: {self.config.template_path}")

            workbook = load_workbook(template_path)
            logger.info(
                "Template loaded successfully",
                extra={"extra_data": {"template_path": self.config.template_path}},
            )
            return workbook

        except TemplateLoadError:
            raise
        except Exception as e:
            logger.error(
                f"Error loading template: {e}",
                extra={"extra_data": {"template_path": self.config.template_path}},
            )
            raise TemplateLoadError(f"Failed to load template: {str(e)}")

    def _fill_header(self, workbook: Any, meta: Meta) -> Any:
        """Fill header cells from meta data"""
        try:
            ws = workbook.active
            meta_dict = meta.model_dump()

            filled_cells = 0
            for field, cell_address in self.header_mapping.items():
                value = meta_dict.get(field)
                if value is not None:
                    ws[cell_address] = value
                    filled_cells += 1
                    logger.debug(
                        f"Header cell filled: {cell_address} = {value}",
                        extra={"extra_data": {"field": field, "cell": cell_address}},
                    )

            logger.info(
                f"Header filled: {filled_cells} cells",
                extra={"extra_data": {"filled_cells": filled_cells}},
            )
            return workbook

        except Exception as e:
            logger.error(f"Error filling header: {e}")
            raise ExcelWriteError(f"Failed to fill header: {str(e)}")

    def _fill_rows(self, workbook: Any, entries: List[Entry]) -> Any:
        """Fill rows from entries data"""
        try:
            ws = workbook.active
            start_row = self.rows_mapping["start_row"]
            columns_mapping = self.rows_mapping["columns"]

            filled_rows = 0
            for index, entry in enumerate(entries):
                row_number = start_row + index
                entry_dict = entry.model_dump()

                for field, column_letter in columns_mapping.items():
                    value = entry_dict.get(field)
                    if value is not None:
                        cell_address = f"{column_letter}{row_number}"
                        cell = ws[cell_address]
                        if field == "percentagem":
                            # API sends 0–100; Excel % format expects a fraction (75 → 0.75 → displays as 75%).
                            cell.value = value / 100
                            cell.number_format = "0%"
                            if value < 100:
                                cell.fill = PatternFill(
                                    start_color=self.config.PERCENTAGE_UNDER_100_FILL,
                                    end_color=self.config.PERCENTAGE_UNDER_100_FILL,
                                    fill_type="solid",
                                )
                        else:
                            cell.value = value
                        logger.debug(
                            f"Row cell filled: {cell_address} = {value}",
                            extra={
                                "extra_data": {
                                    "row": row_number,
                                    "field": field,
                                    "cell": cell_address,
                                }
                            },
                        )

                filled_rows += 1

            logger.info(
                f"Rows filled: {filled_rows} rows",
                extra={"extra_data": {"filled_rows": filled_rows, "total_entries": len(entries)}},
            )
            return workbook

        except Exception as e:
            logger.error(f"Error filling rows: {e}")
            raise ExcelWriteError(f"Failed to fill rows: {str(e)}")

    def _fill_footer(self, workbook: Any, request: GenerateExcelRequest) -> Any:
        """Fill footer cells from mapping (computed date + optional funcionario)."""
        try:
            ws = workbook.active
            funcionario_dict = (
                request.funcionario.model_dump() if request.funcionario else {}
            )
            filled_cells = 0
            for field, cell_address in self.footer_mapping.items():
                if field == "ultimo_dia_util_mes":
                    ws[cell_address] = last_weekday_of_month(
                        date.today().year,
                        request.meta.mes,
                    )
                    filled_cells += 1
                    logger.debug(
                        f"Footer cell filled: {cell_address} = last weekday of month",
                        extra={"extra_data": {"field": field, "cell": cell_address}},
                    )
                    continue
                value = funcionario_dict.get(field)
                if value is not None:
                    ws[cell_address] = value
                    filled_cells += 1
                    logger.debug(
                        f"Footer cell filled: {cell_address} = {value}",
                        extra={"extra_data": {"field": field, "cell": cell_address}},
                    )

            logger.info(
                f"Footer filled: {filled_cells} cells",
                extra={"extra_data": {"filled_cells": filled_cells}},
            )
            return workbook

        except Exception as e:
            logger.error(f"Error filling footer: {e}")
            raise ExcelWriteError(f"Failed to fill footer: {str(e)}")

    def _write_to_stream(self, workbook: Any) -> BytesIO:
        """Write workbook to BytesIO stream"""
        try:
            output_stream = BytesIO()
            workbook.save(output_stream)
            output_stream.seek(0)  # Reset stream position to beginning
            logger.info(
                "Excel written to stream",
                extra={"extra_data": {"stream_size_bytes": output_stream.getbuffer().nbytes}},
            )
            return output_stream

        except Exception as e:
            logger.error(f"Error writing Excel to stream: {e}")
            raise ExcelWriteError(f"Failed to write Excel file: {str(e)}")

    def _apply_calendar_styling(
        self,
        workbook: Any,
        weekend_days: set[int],
        holiday_days: set[int],
    ) -> Any:
        """Highlight column A: holidays (#f6b26b) take priority over weekends."""
        ws = workbook.active
        weekend_fill = PatternFill(
            start_color=self.config.WEEKEND_FILL,
            end_color=self.config.WEEKEND_FILL,
            fill_type="solid",
        )
        holiday_fill = PatternFill(
            start_color=self.config.HOLIDAY_FILL,
            end_color=self.config.HOLIDAY_FILL,
            fill_type="solid",
        )

        for row in range(8, 39):  # Rows 8 to 38
            day_cell = ws[f"A{row}"]
            day_value = day_cell.value
            if day_value is None or not isinstance(day_value, int):
                continue
            if day_value in holiday_days:
                day_cell.fill = holiday_fill
            elif day_value in weekend_days:
                day_cell.fill = weekend_fill

        return workbook

    async def generate(self, request: GenerateExcelRequest) -> BytesIO:
        """
        Generate Excel file from request data
        
        Orchestrates: load_template → fill_header → fill_rows → fill_footer
        → apply_calendar_styling → write_to_stream
        """
        try:
            logger.info(
                "Excel generation started",
                extra={
                    "extra_data": {
                        "meta_fields_provided": len(
                            [v for v in request.meta.model_dump().values() if v is not None]
                        ),
                        "entries_count": len(request.entries),
                    }
                },
            )

            workbook = self.load_template()
            workbook = self._fill_header(workbook, request.meta)
            workbook = self._fill_rows(workbook, request.entries)
            workbook = self._fill_footer(workbook, request)

            current_year = date.today().year
            weekend_days = DateService.get_weekend_days(request.meta.mes, current_year)
            holiday_days = set(request.holidays)
            workbook = self._apply_calendar_styling(
                workbook, weekend_days, holiday_days
            )

            output_stream = self._write_to_stream(workbook)

            logger.info(
                "Excel generation completed successfully",
                extra={"extra_data": {"stream_size_bytes": output_stream.getbuffer().nbytes}},
            )
            return output_stream

        except (TemplateLoadError, ExcelWriteError):
            raise
        except Exception as e:
            logger.error(f"Unexpected error during Excel generation: {e}")
            raise ExcelWriteError(f"Unexpected error: {str(e)}")

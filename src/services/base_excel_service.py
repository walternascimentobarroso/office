# -*- coding: utf-8 -*-
"""Base Excel service with template method pattern for report generation."""

import logging
from abc import ABC, abstractmethod
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.core.config import get_config
from src.core.excel_cells import resolve_writable_cell
from src.core.exceptions import ExcelWriteError, TemplateLoadError
from src.services.date_service import DateService

logger = logging.getLogger(__name__)


def _coerce_day_of_month(value: object) -> int | None:
    """Template cells often store day as float (1.0); normalize to 1–31."""

    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if 1 <= value <= 31 else None
    if isinstance(value, float):
        if not value.is_integer():
            return None
        n = int(value)
        return n if 1 <= n <= 31 else None
    return None


def fill_entry_row_cells(
    ws: Worksheet,
    row_num: int,
    entry: Dict[str, Any],
    columns: Dict[str, str],
    *,
    percentage_under_100_fill: str,
) -> None:
    """Write one entry row; percentagem uses Excel % format and optional fill."""

    for field, col in columns.items():
        value = entry.get(field)
        if value is None:
            continue
        cell_addr = f"{col}{row_num}"
        cell = resolve_writable_cell(ws, cell_addr)
        if field == "percentagem":
            cell.value = value / 100
            cell.number_format = "0%"
            if value < 100:
                cell.fill = PatternFill(
                    start_color=percentage_under_100_fill,
                    end_color=percentage_under_100_fill,
                    fill_type="solid",
                )
        else:
            cell.value = value


class BaseExcelService(ABC):
    """Base service for Excel report generation using template method pattern."""

    def __init__(self, template_path: str) -> None:
        self.template_path = Path(template_path)
        self._config = get_config()

    def generate(self, data: Dict[str, Any], mappings: Dict[str, Any]) -> BytesIO:
        """Template method for Excel generation."""

        logger.info("Generating Excel report from template: %s", self.template_path)
        try:
            wb = load_workbook(self.template_path)
        except OSError as e:
            raise TemplateLoadError(f"Failed to load template: {e}") from e
        except Exception as e:
            raise TemplateLoadError(f"Failed to load template: {e}") from e

        ws = wb.active

        try:
            self.fill_header(ws, data.get("meta", {}), mappings.get("header", {}))
            self.fill_rows(ws, data.get("entries", []), mappings.get("rows", {}))
            self.fill_footer(ws, data, mappings.get("footer", {}))
            self.apply_styles(ws, data)
        except ExcelWriteError:
            raise
        except Exception as e:
            logger.error("Excel generation failed: %s", e, exc_info=True)
            raise ExcelWriteError(f"Failed to generate Excel: {e}") from e

        try:
            output = BytesIO()
            wb.save(output)
            output.seek(0)
        except Exception as e:
            raise ExcelWriteError(f"Failed to write Excel file: {e}") from e

        logger.info("Excel report generated successfully")
        return output

    @abstractmethod
    def fill_header(
        self,
        ws: Worksheet,
        meta: Dict[str, Any],
        header_mappings: Dict[str, str],
    ) -> None:
        """Fill header cells from meta data."""

    @abstractmethod
    def fill_rows(
        self,
        ws: Worksheet,
        entries: List[Dict[str, Any]],
        row_mappings: Dict[str, Any],
    ) -> None:
        """Fill data rows from entries."""

    @abstractmethod
    def fill_footer(
        self,
        ws: Worksheet,
        data: Dict[str, Any],
        footer_mappings: Dict[str, str],
    ) -> None:
        """Fill footer cells."""

    def apply_styles(self, ws: Worksheet, data: Dict[str, Any]) -> None:
        """Highlight column A: holidays take priority over weekends."""

        month_num = DateService.resolve_month(data["meta"]["mes"])
        year = date.today().year
        weekend_days = DateService.get_weekend_days(month_num, year)
        holiday_days = set(data.get("holidays", []))

        weekend_fill = PatternFill(
            start_color=self._config.WEEKEND_FILL,
            end_color=self._config.WEEKEND_FILL,
            fill_type="solid",
        )
        holiday_fill = PatternFill(
            start_color=self._config.HOLIDAY_FILL,
            end_color=self._config.HOLIDAY_FILL,
            fill_type="solid",
        )

        for row in range(8, 39):
            day_cell = resolve_writable_cell(ws, f"A{row}")
            day_num = _coerce_day_of_month(day_cell.value)
            if day_num is None:
                continue
            if day_num in holiday_days:
                day_cell.fill = holiday_fill
            elif day_num in weekend_days:
                day_cell.fill = weekend_fill

# -*- coding: utf-8 -*-
"""Contract tests for POST /generate-excel endpoint"""

import pytest
from fastapi.testclient import TestClient


@pytest.fixture(scope="session")
def setup_test_environment():
    """Setup test environment with temporary template and mappings"""
    import tempfile
    import json
    from pathlib import Path
    from openpyxl import Workbook
    
    with tempfile.TemporaryDirectory() as tmpdir:
        # Create template
        template_path = Path(tmpdir) / "test_template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws['B1'] = None
        ws['D4'] = None
        ws['J4'] = None
        ws['A8'] = 'Day'
        ws['B8'] = 'Description'
        ws['D8'] = 'Location'
        ws['E8'] = 'Start Time'
        ws['J8'] = 'End Time'
        ws['K8'] = 'Percentagem'
        wb.save(str(template_path))
        
        # Create mappings
        header_mapping_path = Path(tmpdir) / "header.json"
        with open(header_mapping_path, 'w') as f:
            json.dump({
                "empresa": "B1",
                "nif": "D4",
                "mes": "J4"
            }, f)
        
        rows_mapping_path = Path(tmpdir) / "rows.json"
        with open(rows_mapping_path, 'w') as f:
            json.dump({
                "start_row": 8,
                "columns": {
                    "day": "A",
                    "description": "B",
                    "location": "D",
                    "start_time": "E",
                    "end_time": "J",
                    "percentagem": "K"
                }
            }, f)
        
        footer_mapping_path = Path(tmpdir) / "footer.json"
        with open(footer_mapping_path, 'w') as f:
            json.dump({
                "nome_completo": "B40",
                "morada": "D40",
                "nif": "J40"
            }, f)
        
        # Set environment variables
        import os
        os.environ["TEMPLATE_PATH"] = str(template_path)
        os.environ["HEADER_MAPPING_PATH"] = str(header_mapping_path)
        os.environ["ROWS_MAPPING_PATH"] = str(rows_mapping_path)
        os.environ["FOOTER_MAPPING_PATH"] = str(footer_mapping_path)
        
        yield tmpdir


@pytest.fixture
def client(setup_test_environment):
    """Create test client for FastAPI app"""
    from src.main import app
    # Reset service singleton
    import src.api.routes.excel as excel_module
    excel_module._service_instance = None
    
    return TestClient(app)


class TestMesValidationContract:
    """Contract tests for mes field validation"""

    def test_mes_required(self, client):
        """Contract: mes field must be present"""
        response = client.post("/generate-excel", json={
            "meta": {"empresa": "Test"},
            "entries": []
        })
        
        assert response.status_code == 422
        response_data = response.json()
        assert "mes" in str(response_data).lower()

    def test_mes_must_be_integer(self, client):
        """Contract: mes must be integer"""
        response = client.post("/generate-excel", json={
            "meta": {"mes": "March"},
            "entries": []
        })
        
        assert response.status_code == 422
        response_data = response.json()
        assert "mes" in str(response_data).lower()

    def test_mes_must_be_1_to_12(self, client):
        """Contract: mes must be between 1 and 12"""
        # Test 0
        response = client.post("/generate-excel", json={
            "meta": {"mes": 0},
            "entries": []
        })
        assert response.status_code == 422
        
        # Test 13
        response = client.post("/generate-excel", json={
            "meta": {"mes": 13},
            "entries": []
        })
        assert response.status_code == 422
        
        # Test valid 6
        response = client.post("/generate-excel", json={
            "meta": {"mes": 6},
            "entries": []
        })
        assert response.status_code == 200
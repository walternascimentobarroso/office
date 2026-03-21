# -*- coding: utf-8 -*-
"""Unit tests for ExcelService"""

import pytest
import tempfile
import json
import os
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from src.core.config import Config
from src.core.exceptions import TemplateLoadError, ExcelWriteError
from src.services.excel_service import ExcelService
from src.core.business_calendar import last_weekday_of_month
from src.models.request import GenerateExcelRequest, Meta, Entry, Funcionario


@pytest.fixture
def temp_config():
    """Create temporary config with files for testing"""
    with tempfile.TemporaryDirectory() as tmpdir:
        # Create template
        template_path = Path(tmpdir) / "template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws['B1'] = None  # empresa
        ws['D4'] = None  # nif
        ws['J4'] = None  # mes
        ws['A8'] = 'Day'
        ws['B8'] = 'Description'
        ws['D8'] = 'Location'
        ws['E8'] = 'Start Time'
        ws['J8'] = 'End Time'
        ws['K8'] = 'Percentagem'
        wb.save(str(template_path))
        
        # Create mappings
        header_mapping_path = Path(tmpdir) / "header.json"
        with open(header_mapping_path, 'w') as f:
            json.dump({
                "empresa": "B1",
                "nif": "D4",
                "mes": "J4"
            }, f)
        
        rows_mapping_path = Path(tmpdir) / "rows.json"
        with open(rows_mapping_path, 'w') as f:
            json.dump({
                "start_row": 9,
                "columns": {
                    "day": "A",
                    "description": "B",
                    "location": "D",
                    "start_time": "E",
                    "end_time": "J",
                    "percentagem": "K"
                }
            }, f)

        footer_mapping_path = Path(tmpdir) / "footer.json"
        with open(footer_mapping_path, 'w') as f:
            json.dump(
                {
                    "ultimo_dia_util_mes": "N47",
                    "nome_completo": "B42",
                    "morada": "B43",
                    "nif": "D44",
                },
                f,
            )
        
        # Set environment variables for test
        old_template = os.environ.get("TEMPLATE_PATH")
        old_header = os.environ.get("HEADER_MAPPING_PATH")
        old_rows = os.environ.get("ROWS_MAPPING_PATH")
        old_footer = os.environ.get("FOOTER_MAPPING_PATH")
        
        try:
            os.environ["TEMPLATE_PATH"] = str(template_path)
            os.environ["HEADER_MAPPING_PATH"] = str(header_mapping_path)
            os.environ["ROWS_MAPPING_PATH"] = str(rows_mapping_path)
            os.environ["FOOTER_MAPPING_PATH"] = str(footer_mapping_path)
            
            yield Config()
        finally:
            # Restore environment
            if old_template:
                os.environ["TEMPLATE_PATH"] = old_template
            else:
                os.environ.pop("TEMPLATE_PATH", None)
            if old_header:
                os.environ["HEADER_MAPPING_PATH"] = old_header
            else:
                os.environ.pop("HEADER_MAPPING_PATH", None)
            if old_rows:
                os.environ["ROWS_MAPPING_PATH"] = old_rows
            else:
                os.environ.pop("ROWS_MAPPING_PATH", None)
            if old_footer:
                os.environ["FOOTER_MAPPING_PATH"] = old_footer
            else:
                os.environ.pop("FOOTER_MAPPING_PATH", None)


class TestExcelServiceLoadTemplate:
    """Tests for loading templates"""

    def test_load_template_success(self, temp_config):
        """Test successfully loading a template"""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        assert workbook is not None
        assert workbook.active is not None

    def test_load_template_not_found(self):
        """Test error when template file not found"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create missing template
            template_path = Path(tmpdir) / "missing.xlsx"
            header_mapping_path = Path(tmpdir) / "header.json"
            rows_mapping_path = Path(tmpdir) / "rows.json"
            footer_mapping_path = Path(tmpdir) / "footer.json"
            
            # Create mapping files but no template
            with open(header_mapping_path, 'w') as f:
                json.dump({"empresa": "B1"}, f)
            with open(rows_mapping_path, 'w') as f:
                json.dump({"start_row": 8, "columns": {}}, f)
            with open(footer_mapping_path, 'w') as f:
                json.dump({"ultimo_dia_util_mes": "N47"}, f)
            
            old_template = os.environ.get("TEMPLATE_PATH")
            old_header = os.environ.get("HEADER_MAPPING_PATH")
            old_rows = os.environ.get("ROWS_MAPPING_PATH")
            old_footer = os.environ.get("FOOTER_MAPPING_PATH")
            
            try:
                os.environ["TEMPLATE_PATH"] = str(template_path)
                os.environ["HEADER_MAPPING_PATH"] = str(header_mapping_path)
                os.environ["ROWS_MAPPING_PATH"] = str(rows_mapping_path)
                os.environ["FOOTER_MAPPING_PATH"] = str(footer_mapping_path)
                
                config = Config()
            except TemplateLoadError:
                # Expected during config initialization
                pass
            finally:
                if old_template:
                    os.environ["TEMPLATE_PATH"] = old_template
                else:
                    os.environ.pop("TEMPLATE_PATH", None)
                if old_header:
                    os.environ["HEADER_MAPPING_PATH"] = old_header
                else:
                    os.environ.pop("HEADER_MAPPING_PATH", None)
                if old_rows:
                    os.environ["ROWS_MAPPING_PATH"] = old_rows
                else:
                    os.environ.pop("ROWS_MAPPING_PATH", None)
                if old_footer:
                    os.environ["FOOTER_MAPPING_PATH"] = old_footer
                else:
                    os.environ.pop("FOOTER_MAPPING_PATH", None)

    def test_load_template_corrupted_file(self):
        """Test error when template file is corrupted"""
        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "corrupted.xlsx"
            header_mapping_path = Path(tmpdir) / "header.json"
            rows_mapping_path = Path(tmpdir) / "rows.json"
            footer_mapping_path = Path(tmpdir) / "footer.json"
            
            # Create corrupted file
            with open(template_path, 'w') as f:
                f.write("not an excel file")
            
            with open(header_mapping_path, 'w') as f:
                json.dump({"empresa": "B1"}, f)
            with open(rows_mapping_path, 'w') as f:
                json.dump({"start_row": 8, "columns": {}}, f)
            with open(footer_mapping_path, 'w') as f:
                json.dump({"ultimo_dia_util_mes": "N47"}, f)
            
            # Don't validate to allow corrupted file
            config = Config.__new__(Config)
            config.template_path = str(template_path)
            config.header_mapping_path = str(header_mapping_path)
            config.rows_mapping_path = str(rows_mapping_path)
            config.footer_mapping_path = str(footer_mapping_path)
            config.header_mapping = {"empresa": "B1"}
            config.rows_mapping = {"start_row": 8, "columns": {}}
            
            service = ExcelService(config)
            with pytest.raises(TemplateLoadError):
                service.load_template()


class TestExcelServiceFillHeader:
    """Tests for filling header rows"""

    def test_fill_header_all_fields(self, temp_config):
        """Test filling header with all fields"""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        ws = workbook.active
        meta = Meta(empresa="Acme Corp", nif="123456", mes="March 2026")
        
        service._fill_header(workbook, meta)
        
        assert ws['B1'].value == "Acme Corp"
        assert ws['D4'].value == "123456"
        assert ws['J4'].value == "March 2026"

    def test_fill_header_missing_fields(self, temp_config):
        """Test filling header with missing fields (none values)"""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        ws = workbook.active
        meta = Meta(empresa="Company", nif=None, mes=None)
        
        service._fill_header(workbook, meta)
        
        assert ws['B1'].value == "Company"
        assert ws['D4'].value is None
        assert ws['J4'].value is None

    def test_fill_header_empty_meta(self, temp_config):
        """Test filling header with empty meta"""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        ws = workbook.active
        meta = Meta()
        
        service._fill_header(workbook, meta)
        
        assert ws['B1'].value is None
        assert ws['D4'].value is None
        assert ws['J4'].value is None


class TestExcelServiceFillRows:
    """Tests for filling data rows"""

    def test_fill_rows_single_entry(self, temp_config):
        """Test filling single entry"""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        ws = workbook.active
        entries = [Entry(day=1, description="Meeting", location="Office", start_time="09:00", end_time="10:00")]
        
        service._fill_rows(workbook, entries)
        
        # Row 8 is header; start_row 9 is the first data row
        assert ws['A9'].value == 1
        assert ws['B9'].value == "Meeting"
        assert ws['D9'].value == "Office"
        assert ws['E9'].value == "09:00"
        assert ws['J9'].value == "10:00"

    def test_fill_rows_percentagem_stores_fraction_for_percent_display(self, temp_config):
        """API sends 0–100; Excel stores fraction so the cell shows e.g. 75%."""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        ws = workbook.active
        entries = [
            Entry(
                day=1,
                description="Meeting",
                location="Office",
                start_time="09:00",
                end_time="10:00",
                percentagem=75,
            )
        ]

        service._fill_rows(workbook, entries)

        assert ws["K9"].value == 0.75
        assert ws["K9"].number_format == "0%"
        assert ws["K9"].fill.start_color.rgb == temp_config.PERCENTAGE_UNDER_100_FILL

    def test_fill_rows_percentagem_100_no_underfill_color(self, temp_config):
        """100% does not apply the under-100 highlight on the Percentagem cell."""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        ws = workbook.active
        entries = [
            Entry(
                day=1,
                description="Meeting",
                location="Office",
                start_time="09:00",
                end_time="10:00",
                percentagem=100,
            )
        ]

        service._fill_rows(workbook, entries)

        assert ws["K9"].value == 1.0
        assert ws["K9"].number_format == "0%"
        assert ws["K9"].fill.start_color.rgb != temp_config.PERCENTAGE_UNDER_100_FILL

    def test_fill_rows_multiple_entries(self, temp_config):
        """Test filling multiple entries"""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        ws = workbook.active
        entries = [
            Entry(day=1, description="Task 1"),
            Entry(day=2, description="Task 2"),
            Entry(day=3, description="Task 3")
        ]
        
        service._fill_rows(workbook, entries)
        
        assert ws['A9'].value == 1
        assert ws['B9'].value == "Task 1"
        assert ws['A10'].value == 2
        assert ws['B10'].value == "Task 2"
        assert ws['A11'].value == 3
        assert ws['B11'].value == "Task 3"

    def test_fill_rows_missing_fields(self, temp_config):
        """Test filling rows with missing fields"""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        ws = workbook.active
        entries = [Entry(day=1, description="Task", location=None, start_time=None, end_time=None)]
        
        service._fill_rows(workbook, entries)
        
        assert ws['A9'].value == 1
        assert ws['B9'].value == "Task"
        assert ws['D9'].value is None
        assert ws['E9'].value is None
        assert ws['J9'].value is None

    def test_fill_rows_empty_entries(self, temp_config):
        """Test filling with empty entries list"""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        ws = workbook.active
        entries = []
        
        service._fill_rows(workbook, entries)
        
        # No changes should be made to data rows
        assert ws['A9'].value is None
        assert ws['B9'].value is None


class TestExcelServiceFillFooter:
    """Tests for footer block (mapping + computed last weekday)."""

    def test_fill_footer_computed_date_and_funcionario(self, temp_config):
        """N47 gets last weekday of meta.mes (same calendar year as generation)."""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        ws = workbook.active
        request = GenerateExcelRequest(
            meta={"mes": 10},
            entries=[],
            funcionario=Funcionario(
                nome_completo="João Silva",
                morada="Rua A, 1",
                nif="123456789",
            ),
        )

        service._fill_footer(workbook, request)

        assert ws["N47"].value == last_weekday_of_month(
            date.today().year,
            request.meta.mes,
        )
        assert ws["B42"].value == "João Silva"
        assert ws["B43"].value == "Rua A, 1"
        assert ws["D44"].value == "123456789"


class TestExcelServiceStreamWriting:
    """Tests for writing workbook to stream"""

    def test_write_to_stream_success(self, temp_config):
        """Test writing workbook to stream"""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        
        stream = service._write_to_stream(workbook)
        
        assert stream is not None
        assert stream.tell() == 0  # Position should be at start
        assert stream.getvalue()  # Should have content
        assert len(stream.getvalue()) > 0

    def test_write_to_stream_returns_bytesio(self, temp_config):
        """Test that write_to_stream returns BytesIO object"""
        from io import BytesIO
        
        service = ExcelService(temp_config)
        workbook = service.load_template()
        
        stream = service._write_to_stream(workbook)
        
        assert isinstance(stream, BytesIO)

    def test_write_to_stream_position_reset(self, temp_config):
        """Test that stream position is reset to 0"""
        service = ExcelService(temp_config)
        workbook = service.load_template()
        
        stream = service._write_to_stream(workbook)
        
        # Should be able to read from beginning
        assert stream.tell() == 0
        content = stream.read(4)
        assert content == b'PK\x03\x04'  # ZIP file header (Excel is ZIP-based)


class TestExcelServiceGenerate:
    """Tests for complete generate workflow"""

    @pytest.mark.asyncio
    async def test_generate_full_workflow(self, temp_config):
        """Test complete generate workflow"""
        service = ExcelService(temp_config)
        request = GenerateExcelRequest(
            meta={"empresa": "Acme", "nif": "123", "mes": 3},
            entries=[
                {"day": 1, "description": "Task 1"},
                {"day": 2, "description": "Task 2"}
            ]
        )
        
        stream = await service.generate(request)
        
        assert stream is not None
        assert stream.tell() == 0
        assert len(stream.getvalue()) > 0

    @pytest.mark.asyncio
    async def test_generate_empty_entries(self, temp_config):
        """Test generate with empty entries"""
        service = ExcelService(temp_config)
        request = GenerateExcelRequest(
            meta={"empresa": "Company"},
            entries=[]
        )
        
        stream = await service.generate(request)
        
        assert stream is not None
        assert len(stream.getvalue()) > 0

    @pytest.mark.asyncio
    async def test_generate_empty_meta(self, temp_config):
        """Test generate with empty meta"""
        service = ExcelService(temp_config)
        request = GenerateExcelRequest(
            meta={},
            entries=[{"day": 1}]
        )
        
        stream = await service.generate(request)
        
        assert stream is not None
        assert len(stream.getvalue()) > 0


class TestExcelServiceWeekendHighlighting:
    """Tests for weekend highlighting functionality"""

    @pytest.mark.asyncio
    async def test_weekend_highlighting_march_31_days(self, temp_config):
        """Test weekend highlighting for month with 31 days"""
        service = ExcelService(temp_config)
        request = GenerateExcelRequest(
            meta={"mes": 3},  # March
            entries=[]
        )
        
        stream = await service.generate(request)
        
        # Load generated Excel
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(stream.getvalue()))
        ws = wb.active
        
        # March 2026 weekends: 1,7,8,14,15,21,22,28,29
        weekend_days = {1, 7, 8, 14, 15, 21, 22, 28, 29}
        fill_color = temp_config.WEEKEND_FILL
        
        for row in range(8, 39):  # Rows 8-38
            day_value = ws[f"A{row}"].value
            if day_value in weekend_days:
                assert ws[f"A{row}"].fill.start_color.rgb == fill_color

    @pytest.mark.asyncio
    async def test_weekend_highlighting_february_28_days(self, temp_config):
        """Test weekend highlighting for February (28 days)"""
        service = ExcelService(temp_config)
        request = GenerateExcelRequest(
            meta={"mes": 2},  # February
            entries=[]
        )
        
        stream = await service.generate(request)
        
        # Load generated Excel
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(stream.getvalue()))
        ws = wb.active
        
        # February 2026 weekends: 1,7,8,14,15,21,22,28
        weekend_days = {1, 7, 8, 14, 15, 21, 22, 28}
        fill_color = temp_config.WEEKEND_FILL
        
        for row in range(8, 36):  # Rows 8-35 (28 days)
            day_value = ws[f"A{row}"].value
            if day_value in weekend_days:
                assert ws[f"A{row}"].fill.start_color.rgb == fill_color

    @pytest.mark.asyncio
    async def test_weekend_highlighting_month_starting_saturday(self, temp_config):
        """Test weekend highlighting when month starts on Saturday"""
        service = ExcelService(temp_config)
        request = GenerateExcelRequest(
            meta={"mes": 9},  # September 2025 starts on Saturday
            entries=[]
        )
        
        stream = await service.generate(request)
        
        # Load generated Excel
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(stream.getvalue()))
        ws = wb.active
        
        # September 2025 weekends: 6,7,13,14,20,21,27,28
        weekend_days = {6, 7, 13, 14, 20, 21, 27, 28}
        fill_color = temp_config.WEEKEND_FILL
        
        for row in range(8, 39):  # Rows 8-38
            day_value = ws[f"A{row}"].value
            if day_value in weekend_days:
                assert ws[f"A{row}"].fill.start_color.rgb == fill_color

    @pytest.mark.asyncio
    async def test_no_data_or_formula_changes(self, temp_config):
        """Test that weekend highlighting doesn't alter data or formulas"""
        service = ExcelService(temp_config)
        request = GenerateExcelRequest(
            meta={"mes": 3, "empresa": "Test Corp"},
            entries=[
                {"day": 1, "description": "Task 1", "location": "Office", "start_time": "09:00", "end_time": "17:00", "percentagem": 100}
            ]
        )
        
        stream = await service.generate(request)
        
        # Load generated Excel
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(stream.getvalue()))
        ws = wb.active
        
        # Check that data is preserved
        assert ws["B1"].value == "Test Corp"  # empresa
        assert ws["A9"].value == 1  # day
        assert ws["B9"].value == "Task 1"  # description
        assert ws["D9"].value == "Office"  # location
        assert ws["E9"].value == "09:00"  # start_time
        assert ws["J9"].value == "17:00"  # end_time
        assert ws["K9"].value == 1.0  # percentagem stored as fraction for Excel %
        assert ws["K9"].fill.start_color.rgb != temp_config.PERCENTAGE_UNDER_100_FILL

    @pytest.mark.asyncio
    async def test_holiday_highlighting_column_a(self, temp_config):
        """Holidays list highlights column A with HOLIDAY_FILL (#f6b26b)."""
        service = ExcelService(temp_config)
        request = GenerateExcelRequest(
            meta={"mes": 3},
            entries=[{"day": 10, "description": "x"}],
            holidays=[10],
        )

        stream = await service.generate(request)

        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(stream.getvalue()))
        ws = wb.active

        assert ws["A9"].value == 10
        assert ws["A9"].fill.start_color.rgb == temp_config.HOLIDAY_FILL

    @pytest.mark.asyncio
    async def test_holiday_priority_over_weekend(self, temp_config):
        """When a day is both weekend and holiday, holiday color wins on column A."""
        service = ExcelService(temp_config)
        # March 15, 2026 is Sunday
        request = GenerateExcelRequest(
            meta={"mes": 3},
            entries=[{"day": 15, "description": "x"}],
            holidays=[15],
        )

        stream = await service.generate(request)

        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(stream.getvalue()))
        ws = wb.active

        assert ws["A9"].value == 15
        assert ws["A9"].fill.start_color.rgb == temp_config.HOLIDAY_FILL
        assert ws["A9"].fill.start_color.rgb != temp_config.WEEKEND_FILL

    @pytest.mark.asyncio
    async def test_holidays_invalid_days_ignored(self, temp_config):
        """Out-of-range and null holiday entries are ignored; valid days still apply."""
        service = ExcelService(temp_config)
        request = GenerateExcelRequest(
            meta={"mes": 3},
            entries=[{"day": 11, "description": "x"}],
            holidays=[None, 0, 32, 11],
        )

        stream = await service.generate(request)

        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(stream.getvalue()))
        ws = wb.active

        assert ws["A9"].fill.start_color.rgb == temp_config.HOLIDAY_FILL

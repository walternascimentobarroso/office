# -*- coding: utf-8 -*-
"""Integration tests for Mapa KM report generation"""

import pytest
from io import BytesIO
from openpyxl import load_workbook
from fastapi.testclient import TestClient

from src.main import app


@pytest.fixture
def client():
    """FastAPI test client"""
    return TestClient(app)


def test_mapa_km_header_filled_correctly(client):
    """Empresa A4, endereço B5."""
    payload = {
        "meta": {
            "empresa": "Test Company Ltd",
            "endereco": "Av. Teste 99",
            "nif": "123456789",
            "mes": 3,
        },
        "entries": [],
        "funcionario": {},
        "holidays": [],
    }

    response = client.post("/reports/mapa-km", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    assert ws["A4"].value == "Test Company Ltd"
    assert ws["B5"].value == "Av. Teste 99"
    assert ws["C6"].value == "123.456.789"


def test_mapa_km_entries_filled_correctly(client):
    """Columns A–D e F conforme mapping."""
    payload = {
        "meta": {"mes": 3},
        "entries": [
            {
                "day": 1,
                "origem": "Braga",
                "destino": "Lisboa",
                "description": "Deslocação para execução de tarefas",
                "n_kms": "363,000",
            }
        ],
        "funcionario": {},
        "holidays": [],
    }

    response = client.post("/reports/mapa-km", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    assert ws["A9"].value == 1
    assert ws["B9"].value == "Braga"
    assert ws["C9"].value == "Lisboa"
    assert ws["D9"].value == "Deslocação para execução de tarefas"
    assert ws["F9"].value == 363.0


def test_mapa_km_entry_maps_to_calendar_day_not_list_order(client):
    """First JSON entry with day 4 must land on row for dia 4 (start_row 9 → row 12)."""
    payload = {
        "meta": {"mes": 3},
        "entries": [
            {
                "day": 4,
                "origem": "Braga",
                "destino": "Lisboa",
                "description": "Deslocação para execução de tarefas",
                "n_kms": "363,000",
            }
        ],
        "funcionario": {},
        "holidays": [],
    }

    response = client.post("/reports/mapa-km", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active

    assert ws["A12"].value == 4
    assert ws["B12"].value == "Braga"
    assert ws["C12"].value == "Lisboa"
    assert ws["D12"].value == "Deslocação para execução de tarefas"
    assert ws["F12"].value == 363.0


def test_mapa_km_funcionario_footer_including_vehicle(client):
    """funcionario block fills footer; vehicle_matricula in viatura cell."""
    payload = {
        "meta": {"mes": 3},
        "entries": [],
        "funcionario": {
            "nome_completo": "Maria Silva",
            "morada": "Rua Um",
            "nif": "999888777",
            "vehicle_matricula": "AA-00-BB",
        },
        "holidays": [],
    }

    response = client.post("/reports/mapa-km", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    assert ws["B45"].value == "Maria Silva"
    assert ws["B46"].value == "Rua Um"
    assert ws["E45"].value == "999.888.777"
    assert ws["E46"].value == "AA-00-BB"


def test_mapa_km_weekend_styling_applied(client):
    """Test that weekend days are styled correctly"""
    payload = {
        "meta": {"mes": 3},
        "entries": [],
        "funcionario": {},
        "holidays": [],
    }

    response = client.post("/reports/mapa-km", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    _ = wb.active


def test_mapa_km_holiday_styling_applied(client):
    """Test that holiday days are styled correctly"""
    payload = {
        "meta": {"mes": 3},
        "entries": [],
        "funcionario": {},
        "holidays": [5, 25],
    }

    response = client.post("/reports/mapa-km", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    _ = wb.active

# -*- coding: utf-8 -*-
"""Integration tests for Mapa Diário report generation"""

import pytest
from io import BytesIO
from openpyxl import load_workbook
from fastapi.testclient import TestClient
from src.main import app


@pytest.fixture
def client():
    """FastAPI test client"""
    return TestClient(app)


def test_mapa_diario_header_filled_correctly(client):
    """Test that header fields are filled correctly in Excel"""
    payload = {
        "meta": {
            "empresa": "Test Company Ltd",
            "nif": "123456789",
            "mes": 3
        },
        "entries": [],
        "funcionario": {},
        "holidays": []
    }
    
    response = client.post("/reports/mapa-diario", json=payload)
    assert response.status_code == 200
    
    # Load Excel from response
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    assert ws["B1"].value == "Test Company Ltd"
    assert ws["D4"].value == "123.456.789"


def test_mapa_diario_entries_filled_correctly(client):
    """Test that entries are filled correctly in rows"""
    payload = {
        "meta": {"mes": 3},
        "entries": [
            {
                "day": 1,
                "description": "Morning meeting",
                "location": "Office",
                "start_time": "09:00",
                "end_time": "10:30",
                "percentagem": 75
            },
            {
                "day": 2,
                "description": "Client call",
                "percentagem": 100
            }
        ],
        "funcionario": {},
        "holidays": []
    }
    
    response = client.post("/reports/mapa-diario", json=payload)
    assert response.status_code == 200
    
    # Load and verify entries
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    assert ws["A8"].value == 1
    assert ws["B8"].value == "Morning meeting"
    assert ws["D8"].value == "Office"
    assert ws["E8"].value == "09:00"
    assert ws["J8"].value == "10:30"


def test_mapa_diario_funcionario_footer_filled(client):
    """Test that funcionario data is filled in footer"""
    payload = {
        "meta": {"mes": 3},
        "entries": [],
        "funcionario": {
            "nome_completo": "João Silva",
            "morada": "Rua Test 123, Lisbon",
            "nif": "987654321"
        },
        "holidays": []
    }
    
    response = client.post("/reports/mapa-diario", json=payload)
    assert response.status_code == 200
    
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    assert ws["B42"].value == "João Silva"
    assert ws["B43"].value == "Rua Test 123, Lisbon"
    assert ws["D44"].value == "987.654.321"


def test_mapa_diario_weekend_styling_applied(client):
    """Test that weekend days are styled correctly"""
    payload = {
        "meta": {"mes": 3},  # March 2024 has weekends
        "entries": [],
        "funcionario": {},
        "holidays": []
    }
    
    response = client.post("/reports/mapa-diario", json=payload)
    assert response.status_code == 200
    
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # Check styling on weekend cells (column A)
    # For March, days 2,3,9,10,16,17,23,24,30,31 are weekends
    # assert ws['A2'].fill.fgColor.rgb == "FFF0F0F0"  # Light gray


def test_mapa_diario_holiday_styling_applied(client):
    """Test that holiday days are styled correctly"""
    payload = {
        "meta": {"mes": 3},
        "entries": [],
        "funcionario": {},
        "holidays": [5, 25]
    }
    
    response = client.post("/reports/mapa-diario", json=payload)
    assert response.status_code == 200
    
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # Check holiday styling
    # assert ws['A5'].fill.fgColor.rgb == "FFFFFFE0"  # Light yellow
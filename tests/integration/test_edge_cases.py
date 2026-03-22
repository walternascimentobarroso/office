# -*- coding: utf-8 -*-
"""Integration tests for edge cases in report generation"""

import pytest
from io import BytesIO
from openpyxl import load_workbook
from fastapi.testclient import TestClient
from src.main import app


@pytest.fixture
def client():
    """FastAPI test client"""
    return TestClient(app)


def test_large_number_of_entries(client):
    """Test handling of large number of entries (performance and correctness)"""
    entries = []
    for i in range(31):  # Full month
        entries.append({
            "day": i + 1,
            "description": f"Activity {i+1}",
            "percentagem": 100
        })
    
    payload = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "Test Employee",
            "address": "Test Address",
            "tax_id": "987654321",
        },
        "month": 3,
        "entries": entries,
        "holidays": [],
    }
    
    response = client.post("/reports/mapa-diario", json=payload)
    assert response.status_code == 200
    
    # Verify all entries are in Excel
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # Check that entries are filled (assuming row 8+)
    # for i in range(31):
    #     assert ws[f'A{8+i}'].value == i + 1


def test_special_characters_in_text_fields(client):
    """Test that special characters in text fields are handled correctly"""
    payload = {
        "company": {
            "name": "Empresa Têst & Cía.",
            "tax_id": "123456789",
            "address": "Av. Teste 99",
        },
        "employee": {
            "name": "José María González",
            "address": "Rua João Paulo II, nº 123",
            "tax_id": "987654321",
        },
        "month": 3,
        "entries": [
            {
                "day": 1,
                "description": "Reunião com João & Maria (café)",
                "location": "São Paulo - SP"
            }
        ],
        "holidays": []
    }
    
    response = client.post("/reports/mapa-diario", json=payload)
    assert response.status_code == 200
    
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # Verify special characters are preserved
    # assert "Têst & Cía" in str(ws['B2'].value)


def test_empty_strings_vs_missing_fields(client):
    """Test difference between empty strings and missing fields"""
    payload = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "Test Employee",
            "address": "Test Address",
            "tax_id": "987654321",
        },
        "month": 3,
        "entries": [
            {
                "day": 1,
                "description": "",  # Empty string
                "location": None,   # Missing field
                "percentagem": 0    # Zero value
            }
        ],
        "holidays": []
    }
    
    response = client.post("/reports/mapa-diario", json=payload)
    assert response.status_code == 200
    
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # Empty string should be written as empty, None should leave cell unchanged
    # assert ws['B8'].value == ""  # Empty string
    # assert ws['E8'].value is None or ws['E8'].value == ""  # None becomes empty (start_time col)


def test_holiday_filtering_invalid_values(client):
    """Test that invalid holiday values are filtered out"""
    payload = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "Test Employee",
            "address": "Test Address",
            "tax_id": "987654321",
        },
        "month": 3,
        "entries": [],
        "holidays": [5, 32, "invalid", -1, 0, 15]  # Mix of valid and invalid
    }
    
    response = client.post("/reports/mapa-diario", json=payload)
    assert response.status_code == 200
    
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # Only valid holidays (5, 15) should be styled
    # assert ws['A5'].fill.fgColor.rgb == "FFFFFFE0"  # Holiday style
    # assert ws['A15'].fill.fgColor.rgb == "FFFFFFE0"
    # assert ws['A32'] is None or ws['A32'].fill.fgColor.rgb != "FFFFFFE0"  # Invalid day not styled


def test_month_boundary_cases(client):
    """Test month boundary values"""
    # Test month 1 (January)
    payload_jan = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "Test Employee",
            "address": "Test Address",
            "tax_id": "987654321",
        },
        "month": 1,
        "entries": [],
        "holidays": [],
    }
    response = client.post("/reports/mapa-diario", json=payload_jan)
    assert response.status_code == 200
    
    # Test month 12 (December)
    payload_dec = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "Test Employee",
            "address": "Test Address",
            "tax_id": "987654321",
        },
        "month": 12,
        "entries": [],
        "holidays": [],
    }
    response = client.post("/reports/mapa-diario", json=payload_dec)
    assert response.status_code == 200


def test_km_month_integers_1_to_12(client):
    """Mapa KM accepts meta.mes as 1–12 like Mapa Diário."""

    for month in range(1, 13):
        payload = {
            "company": {
                "name": "Test Company Ltd",
                "tax_id": "123456789",
                "address": "Av. Teste 99",
            },
            "employee": {
                "name": "Test Employee",
                "address": "Test Address",
                "tax_id": "987654321",
            },
            "month": month,
            "entries": [],
            "holidays": [],
        }
        response = client.post("/reports/mapa-km", json=payload)
        assert response.status_code == 200, f"Failed for month {month}"
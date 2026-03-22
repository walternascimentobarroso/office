# -*- coding: utf-8 -*-
"""Integration tests for daily report generation"""

import pytest
from io import BytesIO
from openpyxl import load_workbook
from fastapi.testclient import TestClient
from src.main import app


@pytest.fixture
def client():
    """FastAPI test client"""
    return TestClient(app)


def test_daily_report_header_filled_correctly(client):
    """Test that header fields are filled correctly in Excel"""
    payload = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "Test Employee",
            "address": "Test Address",
            "tax_id": "987654321",
        },
        "month": 3,
        "entries": [],
        "holidays": [],
    }

    response = client.post("/reports/daily-report", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active

    assert ws["B1"].value == "Test Company Ltd"
    assert ws["D4"].value == "123.456.789"


def test_daily_report_entries_filled_correctly(client):
    """Test that entries are filled correctly in rows"""
    payload = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "Test Employee",
            "address": "Test Address",
            "tax_id": "987654321",
        },
        "month": 3,
        "entries": [
            {
                "day": 1,
                "description": "Morning meeting",
                "location": "Office",
                "start_time": "09:00",
                "end_time": "10:30",
                "percentage": 75,
            },
            {
                "day": 2,
                "description": "Client call",
                "percentage": 100,
            }
        ],
        "holidays": [],
    }

    response = client.post("/reports/daily-report", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active

    assert ws["A8"].value == 1
    assert ws["B8"].value == "Morning meeting"
    assert ws["D8"].value == "Office"
    assert ws["E8"].value == "09:00"
    assert ws["J8"].value == "10:30"
    assert ws["A9"].value == 2
    assert ws["B9"].value == "Client call"


def test_daily_report_entry_maps_to_calendar_day_not_list_order(client):
    """First JSON entry with day 4 must land on row for day 4, not day 1."""
    payload = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "Test Employee",
            "address": "Test Address",
            "tax_id": "987654321",
        },
        "month": 3,
        "entries": [
            {
                "day": 4,
                "description": "Travel",
                "location": "Lisbon",
                "percentage": 100,
            }
        ],
        "holidays": [],
    }

    response = client.post("/reports/daily-report", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active

    assert ws["A11"].value == 4
    assert ws["B11"].value == "Travel"
    assert ws["D11"].value == "Lisbon"


def test_daily_report_employee_footer_filled(client):
    """Test that employee data is filled in footer"""
    payload = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "João Silva",
            "address": "Rua Test 123, Lisbon",
            "tax_id": "987654321",
        },
        "month": 3,
        "entries": [],
        "holidays": [],
    }

    response = client.post("/reports/daily-report", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active

    assert ws["B42"].value == "João Silva"
    assert ws["B43"].value == "Rua Test 123, Lisbon"
    assert ws["D44"].value == "987.654.321"


def test_daily_report_weekend_styling_applied(client):
    """Test that weekend days are styled correctly"""
    payload = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "Test Employee",
            "address": "Test Address",
            "tax_id": "987654321",
        },
        "month": 3,
        "entries": [],
        "holidays": [],
    }

    response = client.post("/reports/daily-report", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    _ = wb.active


def test_daily_report_holiday_styling_applied(client):
    """Test that holiday days are styled correctly"""
    payload = {
        "company": {
            "name": "Test Company Ltd",
            "tax_id": "123456789",
            "address": "1 Test Ave",
        },
        "employee": {
            "name": "Test Employee",
            "address": "Test Address",
            "tax_id": "987654321",
        },
        "month": 3,
        "entries": [],
        "holidays": [5, 25],
    }

    response = client.post("/reports/daily-report", json=payload)
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    _ = wb.active

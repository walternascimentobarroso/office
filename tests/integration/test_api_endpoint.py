# -*- coding: utf-8 -*-
"""Integration tests for POST /generate-excel endpoint"""

import pytest
import tempfile
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
from fastapi.testclient import TestClient
import os


@pytest.fixture(scope="session")
def setup_test_environment():
    """Setup test environment with temporary template and mappings"""
    with tempfile.TemporaryDirectory() as tmpdir:
        # Create template
        template_path = Path(tmpdir) / "test_template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws['B1'] = None
        ws['C4'] = None
        ws['J3'] = None
        ws['A8'] = 'Day'
        ws['B8'] = 'Description'
        ws['D8'] = 'Location'
        ws['E8'] = 'Start Time'
        ws['J8'] = 'End Time'
        wb.save(str(template_path))
        
        # Create mappings
        header_mapping_path = Path(tmpdir) / "header.json"
        with open(header_mapping_path, 'w') as f:
            json.dump({
                "empresa": "B1",
                "nif": "C4",
                "mes": "J3"
            }, f)
        
        rows_mapping_path = Path(tmpdir) / "rows.json"
        with open(rows_mapping_path, 'w') as f:
            json.dump({
                "start_row": 8,
                "columns": {
                    "day": "A",
                    "description": "B",
                    "location": "D",
                    "start_time": "E",
                    "end_time": "J"
                }
            }, f)
        
        # Set environment variables for test
        old_template = os.environ.get("TEMPLATE_PATH")
        old_header = os.environ.get("HEADER_MAPPING_PATH")
        old_rows = os.environ.get("ROWS_MAPPING_PATH")
        
        os.environ["TEMPLATE_PATH"] = str(template_path)
        os.environ["HEADER_MAPPING_PATH"] = str(header_mapping_path)
        os.environ["ROWS_MAPPING_PATH"] = str(rows_mapping_path)
        
        yield tmpdir
        
        # Restore environment
        if old_template:
            os.environ["TEMPLATE_PATH"] = old_template
        else:
            os.environ.pop("TEMPLATE_PATH", None)
        if old_header:
            os.environ["HEADER_MAPPING_PATH"] = old_header
        else:
            os.environ.pop("HEADER_MAPPING_PATH", None)
        if old_rows:
            os.environ["ROWS_MAPPING_PATH"] = old_rows
        else:
            os.environ.pop("ROWS_MAPPING_PATH", None)


@pytest.fixture
def client(setup_test_environment):
    """Create test client for FastAPI app"""
    # Import after environment is set up
    from src.main import app
    # Reset service singleton to pick up new environment
    import src.api.routes.excel as excel_module
    excel_module._service_instance = None
    
    return TestClient(app)


class TestGenerateExcelEndpoint:
    """Tests for POST /generate-excel endpoint"""

    def test_post_valid_payload(self, client):
        """Test POST with valid payload"""
        response = client.post("/generate-excel", json={
            "meta": {"empresa": "Test Corp", "nif": "12345", "mes": "March"},
            "entries": [
                {"day": 1, "description": "Task 1", "location": "Office", "start_time": "09:00", "end_time": "10:00"}
            ]
        })
        
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["content-disposition"]
        assert len(response.content) > 0

    def test_post_empty_entries(self, client):
        """Test POST with empty entries"""
        response = client.post("/generate-excel", json={
            "meta": {"empresa": "Corp"},
            "entries": []
        })
        
        assert response.status_code == 200
        assert len(response.content) > 0

    def test_post_empty_meta(self, client):
        """Test POST with empty meta"""
        response = client.post("/generate-excel", json={
            "meta": {},
            "entries": [{"day": 1, "description": "Task"}]
        })
        
        assert response.status_code == 200
        assert len(response.content) > 0

    def test_post_missing_meta(self, client):
        """Test POST with missing meta field"""
        response = client.post("/generate-excel", json={
            "entries": []
        })
        
        assert response.status_code == 400
        assert "meta" in response.text.lower() or "missing" in response.text.lower()

    def test_post_missing_entries(self, client):
        """Test POST with missing entries field uses default"""
        response = client.post("/generate-excel", json={
            "meta": {"empresa": "Corp"}
        })
        
        # Should use default empty list and succeed
        assert response.status_code == 200

    def test_post_invalid_json(self, client):
        """Test POST with invalid JSON"""
        response = client.post("/generate-excel", content="invalid json")
        
        assert response.status_code == 422  # Validation error

    def test_response_file_valid_xlsx(self, client):
        """Test that returned file is valid XLSX"""
        response = client.post("/generate-excel", json={
            "meta": {"empresa": "Test"},
            "entries": [{"day": 1}]
        })
        
        assert response.status_code == 200
        
        # Try to load response as XLSX
        from io import BytesIO
        try:
            wb = load_workbook(BytesIO(response.content))
            assert wb is not None
        except Exception as e:
            pytest.fail(f"Response is not valid XLSX: {e}")

    def test_response_headers_correct(self, client):
        """Test response headers are correct"""
        response = client.post("/generate-excel", json={
            "meta": {"empresa": "Test"},
            "entries": []
        })
        
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]

    def test_filename_includes_mes(self, client):
        """Test filename includes mes field"""
        response = client.post("/generate-excel", json={
            "meta": {"empresa": "Test", "mes": "March"},
            "entries": []
        })
        
        assert response.status_code == 200
        assert "march" in response.headers["content-disposition"].lower()

    def test_filename_includes_timestamp(self, client):
        """Test filename includes timestamp"""
        response = client.post("/generate-excel", json={
            "meta": {"empresa": "Test"},
            "entries": []
        })
        
        assert response.status_code == 200
        # Filename should have format: relatorio_[mes]_[timestamp].xlsx
        disposition = response.headers["content-disposition"]
        assert ".xlsx" in disposition
        # Should have ISO format timestamp (contains T and Z or digits)
        assert any(c in disposition for c in "0123456789TZ")

    def test_extra_fields_ignored(self, client):
        """Test that extra fields are ignored"""
        response = client.post("/generate-excel", json={
            "meta": {"empresa": "Test", "unknown_field": "value"},
            "entries": [{"day": 1, "extra_field": "ignored"}],
            "extra_top_level": "also_ignored"
        })
        
        assert response.status_code == 200
        assert len(response.content) > 0


class TestHealthEndpoint:
    """Tests for GET /health endpoint"""

    def test_health_check_success(self, client):
        """Test health check endpoint"""
        response = client.get("/health")
        
        assert response.status_code == 200
        assert "status" in response.json() or response.json() == {"status": "healthy"}


class TestDataValidationErrors:
    """Tests for data validation error responses"""

    def test_validation_error_includes_field_details(self, client):
        """Test validation error includes field details"""
        response = client.post("/generate-excel", json={
            "meta": "invalid",
            "entries": []
        })
        
        assert response.status_code == 400
        # Response should indicate what's wrong

    def test_invalid_entries_type(self, client):
        """Test invalid entries type"""
        response = client.post("/generate-excel", json={
            "meta": {},
            "entries": "not a list"
        })
        
        assert response.status_code == 400

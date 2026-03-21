#!/usr/bin/env python3
"""Script to create the Excel template"""

import sys
sys.path.insert(0, '/Users/macbook/projets/office')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = 'Report'

ws['A1'] = 'REPORT'
ws['A1'].font = Font(bold=True, size=14)

ws['A3'] = 'Empresa:'
ws['A4'] = 'NIF:'
ws['A5'] = 'Mes:'

ws['B1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
ws['C4'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
ws['J3'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

ws['A8'] = 'Day'
ws['B8'] = 'Description'
ws['D8'] = 'Location'
ws['E8'] = 'Start Time'
ws['J8'] = 'End Time'

for cell_ref in ['A8', 'B8', 'D8', 'E8', 'J8']:
    ws[cell_ref].font = Font(bold=True, color='FFFFFF')
    ws[cell_ref].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

wb.save('/Users/macbook/projets/office/templates/excel_template2.xlsx')
print('Template created')
